#!/usr/bin/env python3
"""
WACC (Weighted Average Cost of Capital) Report Generator

This script generates a Profit and Loss report for a MeroShare user by:
1. Fetching WACC and share data from user's CSV files
2. Refreshing script details from NEPSE
3. Calculating investment values, current values, and profit/loss
4. Generating formatted Excel and PDF reports
"""

import asyncio
import logging
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import select
from sqlalchemy.orm import selectinload

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.config.settings import config
from src.core.nepse import Meroshare, refresh_all_script_details
from src.database.models import MeroShareUser, Scripts
from src.database.session import get_db

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('logs/wacc_extraction.log')
    ]
)
logger = logging.getLogger(__name__)


async def get_meroshare_user(username: str) -> dict:
    """Fetch MeroShare user from database."""
    logger.info(f"Fetching MeroShare user: {username}")
    async with get_db() as db:
        meroshare_user = (
            await db.execute(
                select(MeroShareUser).filter(MeroShareUser.username == username)
            )
        ).scalars().first()

        if not meroshare_user:
            logger.error(f"MeroShare user not found: {username}")
            raise ValueError(f"User {username} not found in database")

        logger.info(f"Successfully fetched user: {username}")
        return meroshare_user.to_dict()


async def initialize_meroshare(user_data: dict, headless: bool = True) -> Meroshare:
    """Initialize and start MeroShare session."""
    logger.info(f"Initializing MeroShare session for user: {user_data['username']}")
    meroshare = Meroshare(
        headless=headless,
        dp=user_data['dp'],
        password=user_data['password'],
        username=user_data['username']
    )
    await meroshare.start()
    logger.info("MeroShare session started successfully")
    return meroshare


def setup_directories(username: str) -> tuple[Path, Path, Path, Path]:
    """Create necessary directories and return file paths."""
    logger.info("Setting up directories and file paths")
    current_date = pd.Timestamp.now().strftime('%Y-%m-%d')

    base_dir = config.get_user_csv_dir(username)
    os.makedirs(base_dir, exist_ok=True)
    logger.debug(f"Base directory: {base_dir}")

    pnl_dir = Path(base_dir, 'PnL')
    os.makedirs(pnl_dir, exist_ok=True)
    logger.debug(f"P&L directory: {pnl_dir}")

    wacc_rates_csv = Path(base_dir, 'Wacc Rates.csv')
    pnl_xlsx = Path(base_dir, 'P&L.xlsx')
    pnl_pdf = Path(base_dir, 'P&L.pdf')

    # Move existing PDF to archive
    if pnl_pdf.exists():
        archive_path = pnl_dir / f'{current_date}-P&L.pdf'
        logger.info(f"Archiving existing PDF to: {archive_path}")
        subprocess.run(['mv', str(pnl_pdf), str(archive_path)])

    return wacc_rates_csv, pnl_xlsx, pnl_pdf, current_date


def load_source_data(base_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load WACC and shares data from CSV files."""
    logger.info("Loading source data from CSV files")

    wacc_csv = Path(base_dir, 'My Wacc Report.csv')
    shares_csv = Path(base_dir, 'My Shares Values.csv')

    if not wacc_csv.exists():
        logger.error(f"WACC report not found: {wacc_csv}")
        raise FileNotFoundError(f"Missing file: {wacc_csv}")

    if not shares_csv.exists():
        logger.error(f"Shares values not found: {shares_csv}")
        raise FileNotFoundError(f"Missing file: {shares_csv}")

    wacc = pd.read_csv(wacc_csv)
    shares = pd.read_csv(shares_csv)

    logger.info(f"Loaded {len(wacc)} WACC entries and {len(shares)} share entries")
    return wacc, shares


def merge_and_prepare_data(wacc: pd.DataFrame, shares: pd.DataFrame) -> pd.DataFrame:
    """Merge WACC and shares data and prepare initial result DataFrame."""
    logger.info("Merging and preparing data")

    merged_df = pd.merge(wacc, shares, left_on='Scrip Name', right_on='Scrip')
    result = merged_df[['Scrip Name', 'Current Balance', 'WACC Rate']].copy()

    result['WACC Rate'] = result['WACC Rate'].round(2)
    result['Current Balance'] = result['Current Balance'].astype(float)

    result = result.rename(columns={
        'Current Balance': 'Balance',
        'WACC Rate': 'WACC',
        'Scrip Name': 'Scrip'
    })

    logger.info(f"Prepared data for {len(result)} scrips")
    return result


async def fetch_script_details(scrips: list[str]) -> list[Scripts]:
    """Refresh and fetch script details from database."""
    logger.info(f"Refreshing script details for {len(scrips)} scrips")
    await refresh_all_script_details()

    logger.info("Fetching script details from database")
    async with get_db() as db:
        scripts = (
            await db.execute(
                select(Scripts)
                .filter(Scripts.ticker.in_(scrips))
                .options(selectinload(Scripts.script_details))
            )
        ).scalars().all()

    logger.info(f"Fetched {len(scripts)} script details")
    return scripts


def calculate_pnl(result: pd.DataFrame, scripts: list[Scripts]) -> pd.DataFrame:
    """Calculate profit/loss metrics and add totals row."""
    logger.info("Calculating P&L metrics")

    # Extract 52-week high/low data
    high_low = result['Scrip'].apply(
        lambda name: next(
            (script.script_details.week_52_high_low for script in scripts if script.ticker == name),
            None
        )
    )

    result['High'] = high_low.apply(
        lambda x: x.replace(',', '').split(' - ')[0] if isinstance(x, str) else 0
    ).astype(float)

    result['LTP'] = result['Scrip'].apply(
        lambda name: next(
            (script.script_details.last_traded_price for script in scripts if script.ticker == name),
            None
        )
    )

    result['Low'] = high_low.apply(
        lambda x: x.replace(',', '').split(' - ')[1] if isinstance(x, str) else 0
    ).astype(float)

    # Calculate investment and current value
    result['Investment'] = (result['Balance'] * result['WACC']).round(2)
    result['Current Value'] = (result['Balance'] * result['LTP']).round(2)

    # Calculate profit/loss
    result['Profit/Loss'] = (result['Current Value'] - result['Investment']).round(2)
    result['Diff %'] = ((result['Profit/Loss'] / result['Investment']) * 100).round(2)

    # Sort by investment and diff %
    result = result.sort_values(by=['Investment', 'Diff %'], ascending=[False, False])

    # Add totals row
    totals = result[['Investment', 'Current Value', 'Profit/Loss']].sum()
    totals['Scrip'] = 'Total'
    totals['Balance'] = ''
    totals['WACC'] = ''
    totals['LTP'] = ''
    totals['Diff %'] = round((totals['Profit/Loss'] / totals['Investment']) * 100, 2)

    result = pd.concat([result, totals.to_frame().T], ignore_index=True)

    logger.info(f"Total Investment: {totals['Investment']:.2f}")
    logger.info(f"Total Current Value: {totals['Current Value']:.2f}")
    logger.info(f"Total Profit/Loss: {totals['Profit/Loss']:.2f} ({totals['Diff %']:.2f}%)")

    return result


def save_reports(result: pd.DataFrame, wacc_csv: Path, pnl_xlsx: Path):
    """Save CSV and Excel reports."""
    logger.info(f"Saving CSV report to: {wacc_csv}")
    result.to_csv(wacc_csv, index=False)

    logger.info(f"Saving Excel report to: {pnl_xlsx}")
    result.to_excel(pnl_xlsx, index=False)


def format_excel(pnl_xlsx: Path, username: str, current_date: str):
    """Apply formatting to Excel report."""
    logger.info("Applying Excel formatting")

    wb = openpyxl.load_workbook(pnl_xlsx)
    ws = wb.active

    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.85, right=0.25, top=0.5, bottom=0.5)
    ws.page_setup.horizontalCentered = True

    # Add title rows
    ws.insert_rows(1)
    ws.insert_rows(2)

    title = f"[{username}] Profit and Loss Report ({current_date})"
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=14, bold=True)

    # Apply borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = thin_border

    # Set column widths
    column_widths = {
        'A': 10, 'B': 10, 'C': 12, 'D': 12, 'E': 12,
        'F': 12, 'G': 15, 'H': 15, 'I': 12, 'J': 12
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define color fills
    red_fill_1 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    red_fill_2 = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    red_fill_3 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    green_fill_1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    green_fill_2 = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")
    green_fill_3 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    neutral_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Find columns for conditional formatting
    profit_loss_col = None
    current_value_col = None

    for idx, cell in enumerate(ws[3]):
        if cell.value == 'Diff %':
            profit_loss_col = idx + 1
        if cell.value == 'Balance':
            current_value_col = idx + 1

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        profit_loss_value = row[profit_loss_col - 1].value
        current_value_value = row[current_value_col - 1].value

        # Special case for balance = 10
        if current_value_value and current_value_value == 10:
            fill = neutral_fill
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
            continue

        # Color based on profit/loss percentage
        if profit_loss_value:
            if profit_loss_value > 50:
                fill = green_fill_3
            elif profit_loss_value > 20:
                fill = green_fill_2
            elif profit_loss_value > 0:
                fill = green_fill_1
            elif profit_loss_value < -50:
                fill = red_fill_3
            elif profit_loss_value < -20:
                fill = red_fill_2
            elif profit_loss_value < 0:
                fill = red_fill_1
            else:
                fill = neutral_fill

            for cell in row:
                cell.fill = fill

    wb.save(pnl_xlsx)
    logger.info("Excel formatting applied successfully")


async def main(username: str, headless: bool = True):
    """Main execution function."""
    logger.info("=" * 80)
    logger.info(f"Starting WACC extraction for user: {username}")
    logger.info("=" * 80)

    try:
        # Get user data
        user_data = await get_meroshare_user(username)

        # Initialize MeroShare session
        meroshare = await initialize_meroshare(user_data, headless=headless)

        # Setup directories
        wacc_csv, pnl_xlsx, pnl_pdf, current_date = setup_directories(username)

        # Load source data
        base_dir = config.get_user_csv_dir(username)
        wacc, shares = load_source_data(base_dir)

        # Merge and prepare data
        result = merge_and_prepare_data(wacc, shares)

        # Fetch script details
        scripts = await fetch_script_details(result['Scrip'].tolist())

        # Calculate P&L
        result = calculate_pnl(result, scripts)

        # Save reports
        save_reports(result, wacc_csv, pnl_xlsx)

        # Format Excel
        format_excel(pnl_xlsx, username, current_date)

        logger.info("=" * 80)
        logger.info("WACC extraction completed successfully")
        logger.info(f"Reports saved to: {base_dir}")
        logger.info("=" * 80)

    except Exception as e:
        logger.error(f"Error during WACC extraction: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate WACC-based P&L report")
    parser.add_argument(
        "username",
        type=str,
        help="MeroShare username"
    )
    parser.add_argument(
        "--no-headless",
        action="store_true",
        help="Run MeroShare session in non-headless mode (for debugging)"
    )
    args = parser.parse_args()

    asyncio.run(main(args.username, headless=not args.no_headless))
