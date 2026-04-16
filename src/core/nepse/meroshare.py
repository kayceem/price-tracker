import logging
import os
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from playwright.async_api import async_playwright
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from src.config.settings import config
from src.database.models import MeroShareUser, Scripts
from src.database.session import get_db

from .script import refresh_all_script_details


logger = logging.getLogger(__name__)


class Meroshare:
    def __init__(self, dp, username, password, headless=False):
        self.dp = str(dp)
        self.username = username
        self.password = password
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.download_dir = os.path.join(config.get_user_csv_dir(self.username))
        self.urls = {
            "login": "https://meroshare.cdsc.com.np/#/login",
            "portfolio": "https://meroshare.cdsc.com.np/#/portfolio",
            "purchase_source": "https://meroshare.cdsc.com.np/#/purchase",
            "transaction_history": "https://meroshare.cdsc.com.np/#/transaction",
        }

    async def __start(self):
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=False,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
        )
        self.context = await self.browser.new_context(viewport={"width": 1920, "height": 1800})
        self.page = await self.context.new_page()
        await self.page.set_extra_http_headers(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
            }
        )

    async def __close(self):
        await self.browser.close()
        await self.playwright.stop()

    async def login(self):
        await self.page.goto(self.urls["login"])
        await self.page.click("span.select2-selection--single")
        await self.page.fill("input.select2-search__field", self.dp)
        await self.page.keyboard.press("Enter")
        await self.page.fill("input[name='username']", self.username)
        await self.page.fill("input[name='password']", self.password)
        await self.page.click("button[type='submit']")
        await self.page.wait_for_timeout(2000)

    def logout(self):
        pass

    async def calculate_holding_days(self):
        await self.page.goto(self.urls["purchase_source"])
        await self.page.wait_for_timeout(2000)
        holdings_tab = self.page.locator("a.nav-link:has-text('My Holdings')")
        if await holdings_tab.count():
            await holdings_tab.first.click()
            await self.page.wait_for_timeout(2000)

        select_locator = self.page.locator("select#isin")
        await select_locator.wait_for(timeout=5000)

        script_options = await self.page.locator("select#isin option").all()
        scripts = []
        for option in script_options:
            label = (await option.inner_text()).strip()
            value = await option.get_attribute("value")
            if label and value:
                scripts.append({"label": label, "value": value})

        logger.info("Found %s holding scripts to process", len(scripts))

        for idx, script in enumerate(scripts, 1):
            try:
                logger.info("Processing holding days %s/%s: %s", idx, len(scripts), script["label"])
                await select_locator.select_option(script["value"])
                await self.page.wait_for_timeout(500)

                search_button = self.page.locator("button.btn.btn-primary[type='submit']:has-text('Search')")
                await search_button.click()
                await self.page.wait_for_timeout(1500)
                checkboxes = await self.page.locator("input[type='checkbox']").all()
                if not checkboxes:
                    logger.info("No purchase records found for %s", script)
                    await self.page.click("button.btn.btn-default[type='reset']")
                    await self.page.wait_for_timeout(500)
                    continue

                for checkbox in checkboxes[:1]:
                    if await checkbox.is_visible():
                        await checkbox.check()

                await self.page.wait_for_timeout(500)
                proceed_button = self.page.locator("button.btn.btn-primary[type='button']:has-text('Proceed')")
                await proceed_button.click()
                await self.page.wait_for_timeout(1500)

                disclaimer = self.page.locator("input.disclaimer[type='checkbox']")
                if await disclaimer.count() and not await disclaimer.first.is_checked():
                    await disclaimer.first.check()
                await self.page.wait_for_timeout(500)

                update_button = self.page.locator("button.btn.btn-primary[type='submit']:has-text('Update')")
                await update_button.click()
                await self.page.wait_for_timeout(2000)

                toast_container = await self.page.query_selector("div#toast-container")
                if toast_container:
                    toast_text = await toast_container.inner_text()
                    logger.info("Holding days update response for %s: %s", script["label"], toast_text)

                reset_button = self.page.locator("button.btn.btn-default[type='reset']")
                if await reset_button.count():
                    await reset_button.click()
                    await self.page.wait_for_timeout(500)
            except Exception:
                logger.exception("Error processing holding days for %s", script["label"])
                try:
                    reset_button = self.page.locator("button.btn.btn-default[type='reset']")
                    if await reset_button.count():
                        await reset_button.click()
                        await self.page.wait_for_timeout(500)
                except Exception:
                    pass

        logger.info("Completed holding days calculation")
        
    async def calculate_wacc(self):
        await self.page.reload()
        await self.page.wait_for_timeout(2000)
        script_options = await self.page.locator("datalist#browsers option").all()
        scripts = []
        for option in script_options:
            value = await option.get_attribute("value")
            if value:
                scripts.append(value)

        logger.info("Found %s scripts to process", len(scripts))

        for idx, script in enumerate(scripts, 1):
            try:
                logger.info("Processing %s/%s: %s", idx, len(scripts), script)
                await self.page.fill("input#script", script)
                await self.page.wait_for_timeout(500)
                await self.page.click("button.btn.btn-primary[type='submit']")
                await self.page.wait_for_timeout(1500)
                checkboxes = await self.page.locator("input[type='checkbox']").all()
                if not checkboxes:
                    logger.info("No purchase records found for %s", script)
                    await self.page.click("button.btn.btn-default[type='reset']")
                    await self.page.wait_for_timeout(500)
                    continue

                for checkbox in checkboxes:
                    if await checkbox.is_visible():
                        await checkbox.check()

                await self.page.wait_for_timeout(500)
                await self.page.click("button.btn.btn-primary[type='button']:has-text('Proceed')")
                await self.page.wait_for_timeout(1500)
                await self.page.click("input.disclaimer[type='checkbox']")
                await self.page.wait_for_timeout(500)
                await self.page.click("button.btn.btn-primary[type='submit']:has-text('Update')")
                await self.page.wait_for_timeout(2000)

                toast_container = await self.page.query_selector("div#toast-container")
                if toast_container:
                    toast_text = await toast_container.inner_text()
                    logger.info("WACC update response for %s: %s", script, toast_text)

                await self.page.wait_for_timeout(1000)
                await self.page.click("button.btn.btn-default[type='reset']")
                await self.page.wait_for_timeout(500)
            except Exception:
                logger.exception("Error processing WACC for %s", script)
                try:
                    await self.page.click("button.btn.btn-default[type='reset']")
                    await self.page.wait_for_timeout(500)
                except Exception:
                    pass

        logger.info("Completed WACC calculation for all scripts")

    async def fetch_wacc_csv(self):
        while True:
            await self.page.goto(self.urls["purchase_source"])
            await self.page.wait_for_timeout(2000)
            await self.page.click("a.nav-link:has-text('My WACC')")
            await self.page.wait_for_timeout(2000)
            try:
                async with self.page.expect_download() as download_info:
                    await self.page.click("button >> i.msi-download-csv", timeout=5000)
                download = await download_info.value
                download_path = os.path.join(self.download_dir, download.suggested_filename)
                await download.save_as(download_path)
                await self.page.wait_for_timeout(2000)
                return Path(download_path)
            except Exception:
                error_message = await self.page.query_selector("div.fallback-title-message")
                if error_message and "holding" in (await error_message.inner_text()).lower():
                    await self.calculate_holding_days()
                    continue
                if error_message and "wacc" in (await error_message.inner_text()).lower():
                    await self.calculate_wacc()
                    continue
                return None

    async def fetch_protfolio_csv(self):
        await self.page.goto(self.urls["portfolio"])
        await self.page.wait_for_timeout(2000)
        async with self.page.expect_download() as download_info:
            await self.page.click("button >> i.msi-download-csv")
        download = await download_info.value
        download_path = os.path.join(self.download_dir, download.suggested_filename)
        await download.save_as(download_path)
        await self.page.wait_for_timeout(2000)
        return Path(download_path)

    async def fetch_transaction_csv(self):
        await self.page.goto(self.urls["transaction_history"])
        await self.page.wait_for_timeout(2000)
        await self.page.click("input#radio-range")
        await self.page.wait_for_timeout(2000)
        async with self.page.expect_download() as download_info:
            await self.page.click("button >> i.msi-download-csv")
        download = await download_info.value
        download_path = os.path.join(self.download_dir, "history", download.suggested_filename)
        Path(download_path).parent.mkdir(parents=True, exist_ok=True)
        await download.save_as(download_path)
        await self.page.wait_for_timeout(2000)
        return Path(download_path)

    async def fetch_csv(self):
        await self.fetch_protfolio_csv()
        await self.fetch_wacc_csv()
        await self.fetch_transaction_csv()

    async def start(self):
        await self.__start()
        await self.login()
        await self.fetch_csv()
        await self.__close()


class WaccReportGenerator:
    def __init__(self, username: str, headless: bool = True):
        self.username = username
        self.headless = headless

    async def get_meroshare_user(self) -> dict:
        logger.info("Fetching MeroShare user: %s", self.username)
        async with get_db() as db:
            meroshare_user = (
                await db.execute(select(MeroShareUser).filter(MeroShareUser.username == self.username))
            ).scalars().first()
            if not meroshare_user:
                raise ValueError(f"User {self.username} not found in database")
            return meroshare_user.to_dict()

    async def sync_meroshare_csvs(self) -> Meroshare:
        user_data = await self.get_meroshare_user()
        meroshare = Meroshare(
            headless=self.headless,
            dp=user_data["dp"],
            password=user_data["password"],
            username=user_data["username"],
        )
        await meroshare.start()
        return meroshare

    def setup_directories(self) -> tuple[Path, Path, Path, str]:
        current_date = pd.Timestamp.now().strftime("%Y-%m-%d")
        base_dir = config.get_user_csv_dir(self.username)
        os.makedirs(base_dir, exist_ok=True)
        pnl_dir = Path(base_dir, "PnL")
        os.makedirs(pnl_dir, exist_ok=True)
        wacc_rates_csv = Path(base_dir, "Wacc Rates.csv")
        pnl_xlsx = Path(base_dir, "P&L.xlsx")
        pnl_pdf = Path(base_dir, "P&L.pdf")
        if pnl_pdf.exists():
            archive_path = pnl_dir / f"{current_date}-P&L.pdf"
            subprocess.run(["mv", str(pnl_pdf), str(archive_path)], check=False)
        return wacc_rates_csv, pnl_xlsx, pnl_pdf, current_date

    def load_source_data(self, base_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
        wacc_csv = Path(base_dir, "My Wacc Report.csv")
        shares_csv = Path(base_dir, "My Shares Values.csv")
        if not wacc_csv.exists():
            raise FileNotFoundError(f"Missing file: {wacc_csv}")
        if not shares_csv.exists():
            raise FileNotFoundError(f"Missing file: {shares_csv}")
        return pd.read_csv(wacc_csv), pd.read_csv(shares_csv)

    def merge_and_prepare_data(self, wacc: pd.DataFrame, shares: pd.DataFrame) -> pd.DataFrame:
        merged_df = pd.merge(wacc, shares, left_on="Scrip Name", right_on="Scrip")
        result = merged_df[["Scrip Name", "Current Balance", "WACC Rate"]].copy()
        result["WACC Rate"] = result["WACC Rate"].round(2)
        result["Current Balance"] = result["Current Balance"].astype(float)
        return result.rename(
            columns={
                "Current Balance": "Balance",
                "WACC Rate": "WACC",
                "Scrip Name": "Scrip",
            }
        )

    async def fetch_script_details(self, scrips: list[str]) -> list[Scripts]:
        await refresh_all_script_details()
        async with get_db() as db:
            return (
                await db.execute(
                    select(Scripts).filter(Scripts.ticker.in_(scrips)).options(selectinload(Scripts.script_details))
                )
            ).scalars().all()

    def calculate_pnl(self, result: pd.DataFrame, scripts: list[Scripts]) -> pd.DataFrame:
        high_low = result["Scrip"].apply(
            lambda name: next((script.script_details.week_52_high_low for script in scripts if script.ticker == name), None)
        )
        result["High"] = high_low.apply(lambda x: x.replace(",", "").split(" - ")[0] if isinstance(x, str) else 0).astype(float)
        result["LTP"] = result["Scrip"].apply(
            lambda name: next((script.script_details.last_traded_price for script in scripts if script.ticker == name), None)
        )
        result["Low"] = high_low.apply(lambda x: x.replace(",", "").split(" - ")[1] if isinstance(x, str) else 0).astype(float)
        result["Investment"] = (result["Balance"] * result["WACC"]).round(2)
        result["Current Value"] = (result["Balance"] * result["LTP"]).round(2)
        result["Profit/Loss"] = (result["Current Value"] - result["Investment"]).round(2)
        result["Diff %"] = ((result["Profit/Loss"] / result["Investment"]) * 100).round(2)
        result = result.sort_values(by=["Investment", "Diff %"], ascending=[False, False])
        totals = result[["Investment", "Current Value", "Profit/Loss"]].sum()
        totals["Scrip"] = "Total"
        totals["Balance"] = ""
        totals["WACC"] = ""
        totals["LTP"] = ""
        totals["Diff %"] = round((totals["Profit/Loss"] / totals["Investment"]) * 100, 2)
        return pd.concat([result, totals.to_frame().T], ignore_index=True)

    def save_reports(self, result: pd.DataFrame, wacc_csv: Path, pnl_xlsx: Path):
        result.to_csv(wacc_csv, index=False)
        result.to_excel(pnl_xlsx, index=False)

    def format_excel(self, pnl_xlsx: Path, current_date: str):
        wb = openpyxl.load_workbook(pnl_xlsx)
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins = PageMargins(left=0.85, right=0.25, top=0.5, bottom=0.5)
        ws.page_setup.horizontalCentered = True
        ws.insert_rows(1)
        ws.insert_rows(2)
        title = f"[{self.username}] Profit and Loss Report ({current_date})"
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(size=14, bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = thin_border
        column_widths = {
            "A": 10,
            "B": 10,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 15,
            "H": 15,
            "I": 12,
            "J": 12,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        red_fill_1 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        red_fill_2 = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        red_fill_3 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill_1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        green_fill_2 = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")
        green_fill_3 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        neutral_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        profit_loss_col = None
        balance_col = None
        for idx, cell in enumerate(ws[3]):
            if cell.value == "Diff %":
                profit_loss_col = idx + 1
            if cell.value == "Balance":
                balance_col = idx + 1
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            profit_loss_value = row[profit_loss_col - 1].value
            balance_value = row[balance_col - 1].value
            if balance_value and balance_value == 10:
                fill = neutral_fill
            elif profit_loss_value:
                if profit_loss_value > 50:
                    fill = green_fill_3
                elif profit_loss_value > 20:
                    fill = green_fill_2
                elif profit_loss_value > 0:
                    fill = green_fill_1
                elif profit_loss_value < -50:
                    fill = red_fill_3
                elif profit_loss_value < -20:
                    fill = red_fill_2
                elif profit_loss_value < 0:
                    fill = red_fill_1
                else:
                    fill = neutral_fill
            else:
                fill = neutral_fill
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
        wb.save(pnl_xlsx)

    async def generate(self):
        logger.info("Starting WACC extraction for user: %s", self.username)
        await self.sync_meroshare_csvs()
        wacc_csv, pnl_xlsx, _pnl_pdf, current_date = self.setup_directories()
        base_dir = config.get_user_csv_dir(self.username)
        wacc, shares = self.load_source_data(base_dir)
        result = self.merge_and_prepare_data(wacc, shares)
        scripts = await self.fetch_script_details(result["Scrip"].tolist())
        result = self.calculate_pnl(result, scripts)
        self.save_reports(result, wacc_csv, pnl_xlsx)
        self.format_excel(pnl_xlsx, current_date)
        return {
            "base_dir": base_dir,
            "wacc_csv": wacc_csv,
            "pnl_xlsx": pnl_xlsx,
        }
